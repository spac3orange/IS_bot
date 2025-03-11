import gspread
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
import asyncio
from app.core.logger import logger
import openpyxl
from openpyxl.utils import get_column_letter
import os
import re
from tqdm.asyncio import tqdm


class GoogleSheetsClient:
    def __init__(self, credentials_json, sheet_name):
        scopes = [
            'https://www.googleapis.com/auth/spreadsheets',
            'https://www.googleapis.com/auth/drive'
        ]
        credentials = Credentials.from_service_account_file(credentials_json, scopes=scopes)
        self.client = gspread.authorize(credentials)
        self.sheet = self.client.open(sheet_name)

    async def get_all_records_from_all_sheets(self):
        allowed_titles = ['опт раб', 'компьютерка']
        """ Получает все записи со всех вкладок таблицы с минимальным числом API-запросов. """
        all_data = []
        sheets = self.sheet.worksheets()  # Получаем все листы
        problem_titles = [
            "РРЦ ПАЛЕТЫ", "Лист522", "БЕКО РРЦ", "Лист570", "ШИНЫ 3",
            "ШИНЫ КОМПЛЕКТЫ", "РАБОЧКА и НЕ РАБ ТВ", "ТВ и Мониторы",
            "ДАГЕСТАН Рабочка", "НА ХРАНЕНИЕ (холодный склад)", "АСТРАХАНЬ РАБ"
        ]
        for worksheet in tqdm(sheets):
            try:
                await asyncio.sleep(1)
                sheet_name = worksheet.title
                if sheet_name in problem_titles:
                    logger.warning(f'Tab {sheet_name} skipped: tab strucure is not valid ')
                    continue
                logger.info(f'Парсинг вкладки {sheet_name}')
                records = worksheet.get_all_records(expected_headers=['Наименование'])
                all_data.append({sheet_name: records})

            except Exception as e:
                logger.error(f'Ошибка чтения данных из вкладки {sheet_name}: {e}')
                continue

        return all_data

    async def insert_row(self, data):
        """
        Вставляет строку в конец таблицы.
        :param data: Список значений для вставки
        """
        self.sheet.append_row(data)

    async def get_all_records(self):
        """
        Получает все записи из таблицы.
        :return: Список словарей, где ключи - заголовки столбцов
        """
        return self.sheet.get_all_records()

    async def get_row(self, row_number):
        """
        Получает данные конкретной строки.
        :param row_number: Номер строки
        :return: Список значений в строке
        """
        return self.sheet.row_values(row_number)

    async def update_cell(self, row, col, value):
        """
        Обновляет значение конкретной ячейки.
        :param row: Номер строки
        :param col: Номер столбца
        :param value: Новое значение
        """
        self.sheet.update_cell(row, col, value)


    async def download_sheet_as_xlsx(self, output_folder="downloads", filename="main_sheet.xlsx"):
        """
        Загружает всю таблицу и сохраняет её в формате XLSX.
        :param output_folder: Папка, куда сохранять файл
        :param filename: Имя XLSX-файла
        """
        logger.info('Creating sheet backup...')
        if not os.path.exists(output_folder):
            os.makedirs(output_folder)
        try:
            output_path = os.path.join(output_folder, filename)

            workbook = openpyxl.Workbook()
            workbook.remove(workbook.active)  # Удаляем стандартный пустой лист

            for worksheet in tqdm(self.sheet.worksheets()):
                await asyncio.sleep(1)
                sheet_name = worksheet.title
                sanitized_name = re.sub(r'[\/\\:*?"<>|]', '_', sheet_name)  # Заменяем запрещённые символы

                if sheet_name != sanitized_name:
                    logger.info(f'Renamed sheet "{sheet_name}" to "{sanitized_name}" due to invalid characters.')

                logger.info(f'Downloading {sanitized_name}')
                records = worksheet.get_all_values()  # Получаем все значения (включая заголовки)

                if not records:
                    continue  # Пропускаем пустые листы

                ws = workbook.create_sheet(title=sanitized_name)

                for row_idx, row in enumerate(records, start=1):
                    for col_idx, value in enumerate(row, start=1):
                        ws.cell(row=row_idx, column=col_idx, value=value)

                # Автоширина колонок
                for col_idx in range(1, ws.max_column + 1):
                    ws.column_dimensions[get_column_letter(col_idx)].auto_size = True

            workbook.save(output_path)
            logger.info(f'backup created at {output_path}')
            return output_path  # Возвращаем путь к сохранённому файлу
        except Exception as e:
            logger.error(e)




